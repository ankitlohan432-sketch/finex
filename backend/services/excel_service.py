from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from models.user import User, VisitorLog
from models.transactions import Transaction
from datetime import datetime
import os

class ExcelService:
    """Service for exporting data to Excel"""
    
    @staticmethod
    def export_users(db: Session, filename: str = "database/exports/users.xlsx"):
        """Export all users to Excel"""
        os.makedirs("database/exports", exist_ok=True)
        
        users = db.query(User).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Headers
        headers = ["ID", "Email", "Full Name", "Phone", "Role", "Is Active", "Created At", "Last Login"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for user in users:
            ws.append([
                user.id,
                user.email,
                user.full_name,
                user.phone or "",
                user.role,
                "Yes" if user.is_active else "No",
                user.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                user.last_login.strftime("%Y-%m-%d %H:%M:%S") if user.last_login else ""
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(filename)
        return filename
    
    @staticmethod
    def export_visitors(db: Session, filename: str = "database/exports/visitors.xlsx"):
        """Export visitor logs to Excel"""
        os.makedirs("database/exports", exist_ok=True)
        
        visitors = db.query(VisitorLog).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Visitors"
        
        # Headers
        headers = ["ID", "IP Address", "Page", "User Agent", "Visit Time"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data
        for visitor in visitors:
            ws.append([
                visitor.id,
                visitor.ip_address,
                visitor.page,
                visitor.user_agent[:50] + "..." if len(visitor.user_agent) > 50 else visitor.user_agent,
                visitor.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        wb.save(filename)
        return filename
    
    @staticmethod
    def export_transactions(db: Session, filename: str = "database/exports/transactions.xlsx"):
        """Export transactions to Excel"""
        os.makedirs("database/exports", exist_ok=True)
        
        transactions = db.query(Transaction).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Headers
        headers = ["ID", "User ID", "Type", "Status", "Symbol", "Quantity", "Price", "Total Amount", "Created At"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data
        for txn in transactions:
            ws.append([
                txn.id,
                txn.user_id,
                txn.type,
                txn.status,
                txn.symbol or "",
                txn.quantity,
                txn.price_per_unit,
                txn.total_amount,
                txn.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        wb.save(filename)
        return filename

excel_service = ExcelService()
